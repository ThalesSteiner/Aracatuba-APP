import streamlit as st
import time
import pandas as pd
import numpy as np
from streamlit_extras.metric_cards import style_metric_cards
import plotly.express as px
from streamlit_extras.grid import grid
import boto3
import requests
from datetime import datetime
from Aws_pedidos import AWS
import json
import folium
from streamlit_folium import st_folium
from folium.plugins import MarkerCluster
from Funções_APP import Funções
import plotly.express as px
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries


class MultiplasTelas:
    def __init__(self):
        self.tamanho_max_cartela = 133
        self.meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        self.df_pedidos = False
        self.empresas =  ["Nenhuma"] + self.buscar_clientes()


    def Controle_coleta(self):       
        ID_compra = st.text_input("ID compra")
    
            # Definindo variáveis para armazenar os valores inseridos pelo usuário
        data_recebimento = st.date_input("Data de recebimento", format="DD/MM/YYYY")
        mes = data_recebimento.month
        ano = data_recebimento.year
        
        Nome_empresa = st.selectbox("Nome empresa", self.empresas)
        Valor = st.number_input("Valor")
        Detalhe_pedido = st.radio("Algum debito ficou pendente", ["Não", "Sim"])
        if Detalhe_pedido == "Sim":
            Valor_pendente = st.number_input("Valor pendente")    
        Forma_pagamento = st.radio("Selecione forma pagamento:", ["PIX", "Dinheiro", "Débito","Boleto","Cheque"])
        

        if st.button("Confirmar Ordens"):
            if Detalhe_pedido == "Não":
                Valor_pendente = 0
                Status = "Quitado"
                AWS().remover_pedido_nao_pago("Aracatuba Parafusos", str(ID_compra))
            else:
                Status = "Em debito"
            dic = {
                "ID": (ID_compra),
                "Status": str(Status),
                "Data recebimento": str(data_recebimento),
                "Nome da empresa": str(Nome_empresa),
                "Valor": int(Valor),
                "Debito": str(Detalhe_pedido),
                "Valor pendente": str(Valor_pendente),
                "Forma pagamento": str(Forma_pagamento)
            }
            AWS().adicionar_pedido_Controle_Coleta(dic)
            st.success(f"Ordem confirmada, pedido {ID_compra} da empresa {Nome_empresa} no valor de {Valor} enviado")


    def Consulta_pedidos(self):
        tab1, tab2= st.tabs(["Consulta de Débito", "Consulta de Pedido"])

        with tab1:
            st.title("Consulta de Débito")
            debito = st.checkbox("Pesquisar pedidos não pagos")
            if debito:
                buscar = self.buscar_pedidos_nao_pagos()
                Id = st.selectbox("Coloque o ID da consulta", buscar)
                
            else:
                Id = st.text_input("Coloque o ID da consulta")
                
            if st.button("Pesquisar"):
                try:
                    pedido = AWS().buscar_pedido_controle_coleta(Id)
                    mygrid =  grid(3,2, 3, vertical_align="bottom")
                    mygrid.text_input("Status",  value=pedido["Status"])
                    mygrid.text_input("ID",  value=pedido["ID"])
                    mygrid.text_input("Nome da empresa",  value=pedido["Nome da empresa"])
                    mygrid.text_input("Data recebimento",  value=pedido["Data recebimento"])
                    mygrid.text_input("Forma pagamento",  value=pedido["Forma pagamento"])
                    mygrid.text_input("Valor do pedido",  value=float(pedido["Valor"]))
                    mygrid.text_input("Débito",  value=pedido["Debito"])
                    mygrid.text_input("Valor pendente",  value=pedido["Valor pendente"])

                except:
                    st.warning("Pedido não achado")
            
                    
        with tab2:
            st.title("Consulta de Pedido")

            if st.checkbox("Consultar Empresa"):
                empresa = st.selectbox("Selecione a empresa", self.empresas, key="empresa_selectbox")   
                if empresa != "Nenhuma":             
                    pedido = self.buscar_id_pedidos_lojas(empresa)
                    st.session_state.pedidos_empresa = pedido
                    try:
                        if "pedidos_empresa" in st.session_state:
                            Id = st.selectbox("Selecione o ID", st.session_state.pedidos_empresa, key="id_selectbox")
                            if st.button("Buscar Pedido por Empresa"):
                                pedido = AWS().buscar_pedido_ID(Id)
                                Funções().Exibir_pedido(pedido)
                     
                            
                    except:
                        st.warning("Erro ao buscar ID")
            else:
                Id = st.text_input("Coloque o ID da consulta", key="Id2")
                if st.button("Pesquisar", key="Botão2"):
                    pedido = AWS().buscar_pedido_ID(Id)
                    Funções().Exibir_pedido(pedido)

    
    def cadastro_empresa(self):
        tab1, tab2 = st.tabs(["Empresa","Consultar Empresa"])
        with tab1:
            with st.popover("Como cadastrar empresas"):
                st.write("Como cadastrar empresas")
                st.write("https://youtu.be/a4fnYCs0B1I")
                
            st.title("Cadastro de Empresas")
            st.write("Dados Gerais",)
            my_grid = grid(3, 4, [2,1,1], [2,2,1,1],2, vertical_align="bottom")
            #row1
            Nome_nova_empresa = my_grid.text_input("Nome / Razão social")
            Nome_representante = my_grid.text_input("Nome do representante")
            Cpf_cnpj = my_grid.text_input("CPF / CNPJ")
            #row2
            Telefone_contato = my_grid.text_input("Número de celular")
            Telefone_contato2 = my_grid.text_input("Número Fixo")
            Telefone_contato3 = my_grid.text_input("Número de Whatsapp")
            Gmail = my_grid.text_input("Gmail")
            #row3
            Cidade = my_grid.text_input("Cidade")
            Uf = my_grid.text_input("UF")
            Cep = my_grid.text_input("CEP")

            #row4
            Rua = my_grid.text_input("Rua")
            Bairro = my_grid.text_input("Bairro")
            Numero = my_grid.text_input("Numero")
            Complemento = my_grid.text_input("Complemento")

            longitude  = my_grid.text_input("longitude")
            latitude = my_grid.text_input("latitude")
            
            #Rota = st.selectbox("Selecione a Rota:", ["Rota 1", "Rota 2", "Rota 3"])
            Data_cadastro = st.date_input("Data do cadastro", format="DD/MM/YYYY")
        

            if st.button("Confirmar Cadastro"):
                try:
                    dic = {
                        "Nome": Nome_nova_empresa,
                        "Representante": Nome_representante,
                        "Status": "Ativo",
                        "Data cadastro": str(Data_cadastro),
                        "CPF/CNPJ": Cpf_cnpj,
                        "Telefone_contato": Telefone_contato,
                        "Telefone_fixo": Telefone_contato2,
                        "Telefone_whats": Telefone_contato3,
                        "Email": Gmail,
                        "Longitude": longitude,
                        "Latitude": latitude,
                        "Endereco": {
                            "Rua": Rua,
                            "Numero": Numero,
                            "Bairro": Bairro,
                            "Cidade": Cidade,
                            "Uf": Uf,
                            "Cep": Cep,
                            "Complemento": Complemento
                        }
                    }
                    print("TESTANDO")
                    #cadastra o cliente na tabela de endereços geral
                    AWS().cadastro_cliente_endereço(dic)
                    #AWS().adicionar_cliente(Nome_nova_empresa)
                    print("TESTANDO2")
                    AWS().adicionar_cliente_tabela_Cliente_lista(str(Nome_nova_empresa))
                    print("TESTANDO3")
                    AWS().adicionar_loja_tabela_pedidos(Nome_nova_empresa)
                    st.success(f"Cadastro da empresa {Nome_nova_empresa} feito com sucesso!")
                except:
                    st.warning(f"Falha ao cadastrar")
                    
            """with tab2:
            st.title("Cadastro de Expositor")
            st.date_input("Data", format="DD/MM/YYYY")
            my_grid = grid(2, 2, vertical_align="bottom")
            #row1
            empresa = my_grid.selectbox("Nome da empresa", self.empresas)
            my_grid.text_input("Vendedor")
            #row2
            my_grid.text_input("Quantidade")
            my_grid.number_input("Valor")

            if st.button("Confirmar"):
                st.success(f"Expositor da empresa {empresa} vendido com sucesso!")

        with tab3:
            st.title("Cadastro nova rota")
            st.date_input("Data de cadastro", format="DD/MM/YYYY")
            my_grid = grid(1, 1, vertical_align="bottom")
            #row1
            Rota = my_grid.text_input("Nome da Rota")
            #row2
            my_grid.text_area("Descrição da rota")

            if st.button("Cadastar Rota"):
                st.success(f"Nova rota cadastrada com sucesso!")"""
        
        with tab2:
            editar = st.checkbox("Editar empresa")
            st.title("Consulta de Empresa")
            empresa = st.selectbox("Nome da empresa", self.empresas, key="empresa2")
            col1, col2 = st.tabs(["Dados gerais", "Endereço"])

            with col1:
                try:
                    Cadastro = self.buscar_empresa(empresa)
                    if Cadastro["Data cadastro"] == "":
                        Cadastro["Data cadastro"] = "Não cadastrado"
                    mygrid = grid(3, 2, 2, 2, vertical_align="bottom")
                    
                    Nome_nova_empresa = mygrid.text_input("Loja", Cadastro["Nome"])
                    Cpf_cnpj = mygrid.text_input("CPF/CNPJ", Cadastro["CPF/CNPJ"])
                    if editar:
                        Ativo = mygrid.selectbox("Status", ["Ativo", "Inativo"])
                    else:
                        Ativo = mygrid.text_input("Status", Cadastro["Status"])
                    Nome_representante = mygrid.text_input("Representante", Cadastro["Representante"])
                    Telefone_contato = mygrid.text_input("Telefone contato", Cadastro["Telefone_contato"])
                    Data_cadastro = mygrid.text_input("Data do cadastro", Cadastro["Data cadastro"])
                    Gmail = mygrid.text_input("Email", Cadastro["Email"])
                    longitude = mygrid.text_input("Longitude", Cadastro["Longitude"])
                    latitude = mygrid.text_input("Latitude", Cadastro["Latitude"])
                    
                    try:
                        Telefone_contato2 = mygrid.text_input("Telefone fixo", Cadastro["Telefone_fixo"])
                        Telefone_contato3 = mygrid.text_input("Complemento", Cadastro["Telefone_whats"])
                    except:
                        st.warning("Nenhum outro telefone para contato encontrado")
                        
                except:
                    st.warning("Empresa não encontrada")
            
            with col2:
                try:
                    mygrid = grid(3, 2, 2, 1, vertical_align="bottom")
                    Rua = mygrid.text_input("Rua", Cadastro["Endereco"]["Rua"])
                    Numero = mygrid.text_input("Numero", Cadastro["Endereco"]["Numero"])
                    Bairro = mygrid.text_input("Bairro", Cadastro["Endereco"]["Bairro"])
                    Cidade = mygrid.text_input("Cidade", Cadastro["Endereco"]["Cidade"])
                    Uf = mygrid.text_input("Uf", Cadastro["Endereco"]["Uf"])
                    Cep = mygrid.text_input("Cep", Cadastro["Endereco"]["Cep"])
                    Complemento = mygrid.text_input("Complemento", Cadastro["Endereco"]["Complemento"])
                except:
                    st.warning("Empresa não encontrada")
                    
                try:
                    Endereco_anterior = mygrid.text_input("Endereço no antigo sistema", Cadastro["Endereco anterior"])
                except:
                    st.info("Nenhum endereço do antigo sistema foi achado")

            if editar:    
                if st.button("Salvar", key="Botão3"):
                        dic = {
                            "Nome": Nome_nova_empresa,
                            "Representante": Nome_representante,
                            "Status": Ativo,
                            "Data cadastro": Data_cadastro,
                            "CPF/CNPJ": Cpf_cnpj,
                            "Telefone_contato": Telefone_contato,
                            "Telefone_fixo": Telefone_contato2,
                            "Telefone_whats": Telefone_contato3,
                            "Email": Gmail,
                            "Longitude": longitude,
                            "Latitude": latitude,
                            "Endereco": {
                                "Rua": Rua,
                                "Numero": Numero,
                                "Bairro": Bairro,
                                "Cidade": Cidade,
                                "Uf": Uf,
                                "Cep": Cep,
                                "Complemento": Complemento
                            }
                        }
                        AWS().cadastro_cliente_endereço(dic)


    def cadastro_novo_pedido(self):
        st.title("Cadastro de novos pedidos")
        lista = []
        cartela_aco = [129,130,131,132,133]
        for i in range(1, self.tamanho_max_cartela+1):
            lista.append({"Tamanho": f"Cartela {i}", "Quantidade": 0})
        df2 = pd.DataFrame(lista)
        df = st.data_editor(df2, height=500, width=400)
        


        loja = st.selectbox("loja", self.empresas)        
        data = st.date_input("Data do Pedido", format="DD/MM/YYYY")
        venda = st.radio("Forma de Venda", ["Consignado", "Venda"])
        valor_cartela = st.number_input("Valor da Cartela", 2)
        valor_cartela_aço = st.number_input("Valor da Cartela Aço", 3.5)

        lista_parafusos = df["Quantidade"].tolist()
        quantidade_parafusos = sum(lista_parafusos)
        st.title(f"Quantidade de parafusos: {quantidade_parafusos}")
        st.title(f"Valor total do pedido: R${sum(lista_parafusos[:128]) * valor_cartela + (sum(lista_parafusos[128:]) * valor_cartela_aço)}")




        if st.button("Confirmar Cadastro"):
                with st.spinner("Enviando pedido..."):
                    time.sleep(6)
                    df = df[df['Quantidade'] != 0]
                    #df["Tamanho"] = [cartela.split(" ")[1] for cartela in df["Tamanho"].tolist()]
                    pedido = dict(zip(df['Tamanho'], df['Quantidade']))
                    quantidade_total_parafusos = sum(df["Quantidade"].tolist())
                    
                    if pedido == {'Tamanho': {}, 'Quantidade': {}}:
                        st.warning("Pedido vazio")
                    elif quantidade_total_parafusos < 50:
                        self.Aviso_pedido(f"Pedido total com: {quantidade_total_parafusos} parafusos está abaixo do mínimo de 50")
                    elif loja == "Nenhuma":
                        self.Aviso_pedido(f"Por favor selecione uma loja")
                    else:
                        try:
                            quantidade_parafusos = sum(df['Quantidade'].tolist())
                            
                            quantidade_p_aco = []
                            for N_cartela, N_parafuso in pedido.items():
                                numero_cartela = int(N_cartela.split(" ")[1])
                                if numero_cartela in cartela_aco:
                                    quantidade_p_aco.append(int(N_parafuso))
                            
                            Id = AWS().Gerar_novo_id()
                            
                            dic = {"ID": str(Id),
                                "Loja": loja,
                                "Data": str(data),
                                "Hora": str(self.Data_hora()),
                                "Tipo de Venda": venda,
                                "Valor da cartela": str(valor_cartela),
                                "valor da cartela aço":str(valor_cartela_aço),
                                "Pedidos": pedido
                                }

                            valor_pedido = str((((int(quantidade_parafusos) - sum(quantidade_p_aco)) * float(valor_cartela)) + (int(sum(quantidade_p_aco)) *  float(valor_cartela_aço))))
                            controle_coleta = {
                                "ID": (str(Id)),
                                "Status": str("Em debito"),
                                "Data recebimento": str(data),
                                "Nome da empresa": str(loja),
                                "Valor": valor_pedido,
                                "Debito": "Sim",
                                "Valor pendente": valor_pedido,
                                "Forma pagamento": "Não declarada"}
                            
                            data = f"{self.meses[int(data.month)-1]}-{data.year}"
                            AWS().adicionar_pedido_tabela_pedidos(loja,Id)
                            AWS().adicionar_pedido_tabela_pedidos_gerais(data,Id,dic)
                            AWS().adicionar_pedido_tabela_pedidosID(dic)
                            AWS().adicionar_pedido_Controle_Coleta(controle_coleta)
                            
                            AWS().adicionar_pedido_nao_pago("Aracatuba Parafusos", str(Id))
                            
                            self.Aviso_pedido(f"""O pedido ID: {Id} da empresa {loja} foi confirmado, 
                                            Total de parafusos: {quantidade_total_parafusos}, 
                                            Preço total do pedido: {valor_pedido}""")
                            
                        except:
                            st.warning("ERRO AO CADASTRAR")
   

    def Separar_pedido(self):
        
        st.title("Separar Pedido")
        my_grid = grid(3, vertical_align="bottom")
        data_nota = st.date_input("Data de emissão", format="DD/MM/YYYY")
        Id1 = my_grid.text_input("ID")
        Id2 = my_grid.text_input("ID", key="2")
        Id3 = my_grid.text_input("ID", key="3")

        if st.button("Pesquisar"):
            # Carregar o modelo de referência
            df = pd.read_excel("Planilha_referencia.xlsx")

            lista_id = [Id1, Id2, Id3]
            df.iloc[1, 1] = lista_id[0]
            df.iloc[1, 10] = lista_id[1]
            df.iloc[1, 19] = lista_id[2]
            lista_index = [[1, 3, 5, 7], [10, 12, 14, 16], [19, 21, 23, 25]]
            lojas = []
            preço_p = []
            preço_p_aco = []
            quantidade_p = []
            quantidade_p_aco = []
            i = 0

            for id in lista_id:
                if id != '':
                    pedido = AWS().buscar_pedido_ID(id)
                    lojas.append(pedido["Loja"])
                    preço_p.append(float(pedido["Valor da cartela"]))
                    preço_p_aco.append(float(pedido["valor da cartela aço"]))

                    pedido_dic = dict(pedido["Pedidos"])
                    lista_cartela = [0 for _ in range(self.tamanho_max_cartela)]
                    for cartela, quantidade in pedido_dic.items():
                        lista_cartela[int(cartela.split(" ")[1]) - 1] = int(quantidade)

                    #Colunas com o somatorio
                    lista_1 = lista_cartela[:34] + [sum(lista_cartela[:34])]
                    lista_2 = lista_cartela[34:68] + [sum(lista_cartela[34:68])]
                    lista_3 = lista_cartela[68:102] + [sum(lista_cartela[68:102])]
                    lista_4 = lista_cartela[102:] + [sum(lista_cartela[102:])]

                    quantidade_p_aco.append(sum(lista_cartela[128:]))
                    
                    df.iloc[4:39, lista_index[i][0]] = lista_1
                    df.iloc[4:39, lista_index[i][1]] = lista_2
                    df.iloc[4:39, lista_index[i][2]] = lista_3

                    df.iloc[4:35, lista_index[i][3]] = lista_4[:-1]

                    # Adiciona a soma na última coluna que é diferente das demais
                    df.iloc[38, lista_index[i][3]] = lista_4[-1]

                    quantidade_p.append(sum(lista_cartela))

                    i += 1
                else:
                    lojas.append("")
                    preço_p.append("")
                    preço_p_aco.append("")
                    quantidade_p.append("")
                    i += 1
            
            df.iloc[2, 1] = lojas[0]
            df.iloc[2, 10] = lojas[1]
            df.iloc[2, 19] = lojas[2]

            df.iloc[42, 0] = f"{preço_p[0]} / {preço_p_aco[0]}"
            df.iloc[42, 9] = f"{preço_p[1]} / {preço_p_aco[1]}"
            df.iloc[42, 18] = f"{preço_p[2]} / {preço_p_aco[2]}"

            df.iloc[42, 3] = quantidade_p[0]
            df.iloc[42, 12] = quantidade_p[1]
            df.iloc[42, 21] = quantidade_p[2]

            try:
                df.iloc[42, 6] = ((float(quantidade_p[0]) - float(quantidade_p_aco[0])) * float(preço_p[0])) + (float(quantidade_p_aco[0]) * float(preço_p_aco[0]))
            except:
                pass

            try:
                df.iloc[42, 15] = ((float(quantidade_p[1]) - float(quantidade_p_aco[1])) * float(preço_p[1])) + (float(quantidade_p_aco[1]) * float(preço_p_aco[1]))
            except:
                pass

            try:
                df.iloc[42, 24] = ((float(quantidade_p[2]) - float(quantidade_p_aco[2])) * float(preço_p[2])) + (float(quantidade_p_aco[2]) * float(preço_p_aco[2]))
            except:
                pass

            df.replace(0 , "", inplace=True)
            output_file = "Modelo_temporario.xlsx"
            
            # Salvar o DataFrame modificado em um novo arquivo Excel temporário
            df.to_excel(output_file, index=False)

            # Carregar a planilha existente com openpyxl
            workbook = load_workbook('Planilha_referencia.xlsx')
            sheet = workbook.active

            # Abrir o arquivo Excel salvo temporariamente com o DataFrame modificado
            df_modified = pd.read_excel(output_file)

            # Aplicar as modificações do DataFrame na planilha carregada
            for r_idx, row in df_modified.iterrows():
                for c_idx, value in enumerate(row):
                    cell = sheet.cell(row=r_idx + 1, column=c_idx + 1)

                    # Verifica se a célula faz parte de uma região mesclada
                    if cell.coordinate in sheet.merged_cells:
                        # Obtém a faixa mesclada (ex: A1:B2)
                        for merged_range in sheet.merged_cells.ranges:
                            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                            if min_row == r_idx + 1 and min_col == c_idx + 1:
                                # Somente a célula superior esquerda deve ser modificada
                                cell.value = value
                                break
                    else:
                        # Caso a célula não faça parte de uma região mesclada
                        cell.value = value

            # Salvar o workbook com as modificações
            workbook.save('Entrega_Final.xlsx')
            
            # Fechar o workbook
            workbook.close()
            
            with open("Entrega_Final.xlsx", "rb") as file:
                # Criar o botão de download para baixar o arquivo Excel exatamente como ele está
                st.download_button(
                    label="Baixar planilha",
                    data=file,
                    file_name=f"Lista dos pedidos ID: {Id1} {Id2} {Id3} .xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                        
       

    def Data_hora(self):
        try:
            response = requests.get('http://worldtimeapi.org/api/timezone/America/Sao_Paulo')
            # Extrair a data e hora atuais da resposta
            data_hora_atual = response.json()['datetime']
            data_formatada = data_hora_atual[0:10]
            Hora = data_hora_atual[11:19]
            #st.write(data_formatada, Hora)
            return f"{Hora}"
        except:
            hora_sistema = datetime.now()
            hora_sistema_formatada = hora_sistema.strftime('%Y-%m-%d %H:%M:%S')
            #st.write("Data e hora do sistema (formatada):", hora_sistema_formatada)
            return f"{hora_sistema_formatada}"


    @st.cache_data
    def gerar_dado(_self, empresas):
        pedidos = []

        anos = [f"{2017+i}" for i in range(8)]
        meses_do_ano = [
            "Janeiro", "Fevereiro", "Março", "Abril",
            "Maio", "Junho", "Julho", "Agosto",
            "Setembro", "Outubro", "Novembro", "Dezembro"
        ]
        lojas = empresas

        for ano in anos:
            for mes in meses_do_ano:
                for loja in lojas[1:]:
                    repeticao = np.random.randint(25, 80)
                    dic = {"Loja": loja,
                            "Data": f"{mes}-{ano}"}
                    dic2 = {}
                    for z in range(1, repeticao):
                        parafuso = str(np.random.randint(1, 129))
                        quantidade = str(np.random.randint(1, 16))
                        dic2[parafuso] = quantidade
                    dic["Cartela"] = dic2
                    pedidos.append(dic)

        return pedidos



    def Dashboard(self):
        pedidos = self.gerar_dado(self.empresas)
        #st.json(pedidos)
        st.title("Pedidos")
        tabela_completa = st.toggle('Tabela Completa')
        my_grid = grid(2,3, vertical_align="bottom")
        Cidade = my_grid.multiselect("Escolha a Cidade", ["todos", "Rio de janeiro", "São joão de meriti", "Nova Iguaçu"])
        Bairro = my_grid.multiselect("Escolha a Bairro", ["todos", "Piedade", "Meier", "Madureira"])
        #Bairro = my_grid.multiselect("Escolha a Bairro", ["todos", "Rota1", "Rota2", "Rota3"])
        
        loja = my_grid.multiselect("Escolha as Lojas", ["Todos"] + self.empresas)
        mês = my_grid.multiselect("Escolha o Mês", ["Todos"] + self.meses)
        ano = my_grid.multiselect("Escolha o Ano", ["Todos"] + [f"{2017+i}" for i in range(8)])

        
        if st.button("Pesquisar"):
            
            if "Todos" not in loja:
                pedidos = [pedido for pedido in pedidos if pedido['Loja'] in loja]
            if "Todos" not in mês:
                pedidos = [pedido for pedido in pedidos if pedido['Data'].split("-")[0] in mês]
            if "Todos" not in ano:
                pedidos = [pedido for pedido in pedidos if pedido['Data'].split("-")[1] in ano]
                
            # Inicializa os dicionários
            itens = self.tamanho_max_cartela
            dic = {str(i): 0 for i in range(1, itens + 1)}
            datas_unicas = {}
            Anos_unicas = {}
            Mês_unicas = {}
            Lista_lojas = {}
            Venda_total_parafusos = {}

            # Itera sobre os pedidos
            for pedido in pedidos:
                # Atualiza contagem de parafusos
                for cartela, quantidade in pedido["Cartela"].items():
                    dic[cartela] += int(quantidade)
                    if cartela in Venda_total_parafusos:
                        Venda_total_parafusos[cartela] += int(quantidade)
                    else:
                        Venda_total_parafusos[cartela] = int(quantidade)
                
                # Calcula a soma das cartelas
                cartelas = [int(cartela) for cartela in pedido['Cartela']]
                soma_cartelas = sum(cartelas)
                
                # Atualiza a soma para a data correspondente
                data = pedido["Data"]
                if data in datas_unicas:
                    datas_unicas[data] += soma_cartelas
                else:
                    datas_unicas[data] = soma_cartelas

                # Atualiza a soma para o ano correspondente
                ano = data.split("-")[1]
                if ano in Anos_unicas:
                    Anos_unicas[ano] += soma_cartelas
                else:
                    Anos_unicas[ano] = soma_cartelas

                # Atualiza a soma para o mês correspondente
                mes = data.split("-")[0]
                if mes in Mês_unicas:
                    Mês_unicas[mes] += soma_cartelas
                else:
                    Mês_unicas[mes] = soma_cartelas
                
                # Atualiza a soma para a loja correspondente
                loja = pedido["Loja"]
                if loja in Lista_lojas:
                    Lista_lojas[loja] += soma_cartelas
                else:
                    Lista_lojas[loja] = soma_cartelas

            # Calcula a média de cada parafuso
            total_pedidos = len(pedidos)
            Media_cartela = {cartela: round(quantidade / total_pedidos, 2) for cartela, quantidade in dic.items()}
            
            # Converter dados para DataFrame
            df_cartelas = pd.DataFrame(list(dic.items()), columns=['Cartela', 'Total'])
            df_datas = pd.DataFrame(list(datas_unicas.items()), columns=['Data', 'Total'])
            df_anos = pd.DataFrame(list(Anos_unicas.items()), columns=['Ano', 'Total'])
            df_meses = pd.DataFrame(list(Mês_unicas.items()), columns=['Mês', 'Total'])
            df_lojas = pd.DataFrame(list(Lista_lojas.items()), columns=['Loja', 'Total'])

            # Criar gráficos
            fig_cartelas = px.bar(df_cartelas, x='Cartela', y='Total', title='Total de Parafusos por Cartela')
            fig_datas = px.line(df_datas, x='Data', y='Total', title='Total de Cartelas por Data')
            fig_anos = px.bar(df_anos, x='Ano', y='Total', title='Total de Cartelas por Ano')
            fig_meses = px.bar(df_meses, x='Mês', y='Total', title='Total de Cartelas por Mês')
            fig_lojas = px.bar(df_lojas, x='Loja', y='Total', title='Total de Cartelas por Loja')

            
            my_grid = grid(4, vertical_align="bottom")
            
            # Mostrar gráficos no Streamlit
            st.title('Análise de Pedidos Simulados')

            st.subheader('Total de Parafusos por Cartela')
            st.plotly_chart(fig_cartelas)

            st.subheader('Total de Cartelas por Data')
            st.plotly_chart(fig_datas)

            st.subheader('Total de Cartelas por Ano')
            st.plotly_chart(fig_anos)

            st.subheader('Total de Cartelas por Mês')
            st.plotly_chart(fig_meses)

            st.subheader('Total de Cartelas por Loja')
            st.plotly_chart(fig_lojas)
            
            
            st.dataframe(df_cartelas)
            st.dataframe(df_datas)
            st.dataframe(df_anos)
            st.dataframe(df_meses)
            st.dataframe(df_lojas)
        
    
    def atualizar_estoque(self, df, estoque, Id=None):
        dic_estoque = df.to_dict()
        Cartela_dataframe = dic_estoque["Cartela"]
        Quantidade_dataframe = dic_estoque["Quantidade"]
        dic_estoque = {cartela: int(quantidade) for cartela, quantidade in zip(Cartela_dataframe.values(), Quantidade_dataframe.values())}
        
        if Id:
            pedido = AWS().buscar_pedido_ID(Id)
            for cartela, quantidade in pedido["Pedidos"].items():
                dic_estoque[cartela] = dic_estoque.get(cartela, 0) - quantidade

        AWS().Adicionar_estoque_cartela(estoque, dic_estoque)
        self.tabela_estoque_formatada.clear()
        st.success("Estoque atualizado")
        time.sleep(2)
        st.rerun()
    
        
    
    def Estoque(self):
        tab1, tab2 = st.tabs(["Estoque cartela", "Estoque caixa"])
        df = self.tabela_estoque_formatada("Parafuso Cartela")
        df2 = self.tabela_estoque_formatada("Parafuso Caixa")

        
        with tab1:
            if st.checkbox("Editar Estoque"):
                dataframe = st.data_editor(df, width=500, height=700)
                if st.button("Salvar estoque"):
                    self.atualizar_estoque(dataframe, "Parafuso Cartela")
            else:
                Id = st.text_input("Buscar ID")
                st.title("Estoque cartela")  
                dataframe = st.dataframe(df, width=500, height=700)
                if st.button("Salvar estoque"):
                    self.atualizar_estoque(df, "Parafuso Cartela", Id)
        
        with tab2:
            if st.checkbox("Editar Estoque", key="Estoque2"):
                dataframe = st.data_editor(df2, width=500, height=700, key="Dataframe2")
                if st.button("Salvar estoque", key="Botão2"):
                    self.atualizar_estoque(dataframe, "Parafuso Caixa")
            else:
                Id = st.text_input("Buscar ID", key="Id2")
                st.title("Estoque cartela")  
                dataframe2 = st.dataframe(df2, width=500, height=700)
                if st.button("Salvar estoque", key="Botão3"):
                    self.atualizar_estoque(df2, "Parafuso Caixa", Id)
    
      
    def Rotas(self):
        Cadastro_empresas = self.buscar_cadastro_empresas()

        latitudes = []
        longitudes = []
        lojas = []
        Status = []
        endereços = []
        bairros = []
        cidades = []

        for empresa in Cadastro_empresas:
            try:
                longitudes.append(float(empresa["Longitude"]))
                latitudes.append(float(empresa["Latitude"]))
                lojas.append(empresa["Nome"])
                Status.append(empresa["Status"])
                endereços.append(empresa["Endereco"])
                bairros.append(empresa["Endereco"]["Bairro"])
                cidades.append(empresa["Endereco"]["Cidade"])
            except:
                pass

        if not latitudes or not longitudes:
            st.write("Nenhuma coordenada válida encontrada.")
            return

        lojas.insert(-1, "Todas")
        bairros.insert(-1, "Todos")
        cidades.insert(-1, "Todas")
        
        
        mygrid =  grid(2, 2, vertical_align="bottom")
        # Adicionar filtros
        status_selecionados = mygrid.multiselect("Selecione o status das lojas", options=set(Status), default=["Ativo"])
        lojas_selecionadas = mygrid.multiselect("Selecione as lojas", options=set(lojas), default=["Todas"])
        bairros_selecionados = mygrid.multiselect("Selecione os bairros", options=set(bairros), default=["Todos"])
        cidades_selecionadas = mygrid.multiselect("Selecione as cidades", options=set(cidades), default=["Todas"])


        # Filtrar dados com base nas seleções
        filtro = [
            (lat, lon, loja, statu, endereco2)
            for lat, lon, loja, statu, endereco2 in zip(latitudes, longitudes, lojas, Status, endereços)
            if (statu in status_selecionados) and
            (lojas_selecionadas == ["Todas"] or loja in lojas_selecionadas) and
            (bairros_selecionados == ["Todos"] or endereco2["Bairro"] in bairros_selecionados) and
            (cidades_selecionadas == ["Todas"] or endereco2["Cidade"] in cidades_selecionadas)
        ]

        if not filtro:
            st.write("Nenhuma loja encontrada com os filtros selecionados.")
            return

        mapa = folium.Map(location=[-22.832113794507347, -43.346853465267536], zoom_start=13)
        marker_cluster = MarkerCluster().add_to(mapa)

        for lat, lon, loja, statu, endereco2 in filtro:
            try:
                rua = endereco2['Rua']
                numero = endereco2['Numero']
                bairro = endereco2['Bairro']
                cidade = endereco2['Cidade']
                cep = endereco2['Cep']

                endereco = f"{rua} {numero} {bairro} {cidade} {cep}"
            except:
                endereco = endereco2

            if statu == "Ativo":
                folium.Marker(
                    [lat, lon],
                    popup=folium.Popup(
                        f"<b>{loja}</b><br>{endereco}<br>", 
                        max_width=300
                    ),
                    icon=folium.Icon(icon="cloud")
                ).add_to(mapa)
            elif statu == "Inativo":
                folium.Marker(
                    [lat, lon],
                    popup=folium.Popup(
                        f"<b>{loja}</b><br>{endereco}<br>", 
                        max_width=300
                    ),
                    icon=folium.Icon(icon="remove", color="red")
                ).add_to(mapa)
            else:
                folium.Marker(
                    [lat, lon],
                    popup=folium.Popup(f"{statu} {loja} {endereco}", max_width=300),
                    icon=folium.Icon(icon="remove", color="purple")
                ).add_to(mapa)

        folium.Marker(
            [-22.832113794507347, -43.346853465267536],
            popup="Araçatuba Material",
            icon=folium.Icon(icon="home", color="orange")
        ).add_to(mapa)

        st_folium(mapa, width=1800, height=800)
    
    
    def Rotas2(self):
        df = pd.read_csv("Lojas_RJ_csv.csv")
        bairros = pd.read_csv("Bairros.csv")["Lista bairros"].tolist()
        
        
        mygrid =  grid(1, vertical_align="bottom")
        
        status_selecionados = mygrid.multiselect("Selecione o bairro", options=set(bairros))
        
        if status_selecionados == "":
            pass
        else:
            # Criar uma condição para verificar se algum dos bairros está contido na string
            filtro = df['Endereços'].apply(lambda x: any(bairro in x for bairro in status_selecionados))
            df_filtrado = df[filtro]
        
        long_media = np.mean(df_filtrado["Longitude"].astype(float).tolist())
        lat_media = np.mean(df_filtrado["Latitude"].astype(float).tolist())
        try:
            mapa = folium.Map(location=[lat_media, long_media], zoom_start=14)
            marker_cluster = MarkerCluster().add_to(mapa)
        except:
            mapa = folium.Map(location=[-22.832113794507347, -43.346853465267536], zoom_start=14)
            marker_cluster = MarkerCluster().add_to(mapa)
        for N_loja, empresa, endereço, link_maps, tipo, lat, lon in df_filtrado.iloc[:, -7:].values:    
            try:
                endereco = f"{endereço}"
            except:
                endereco = "Endereço não encontrado"

            try:
                folium.Marker(
                    [lat, lon],
                    popup=folium.Popup(
                        f"<b>{empresa}</b><br>{endereco}<br><a href='{link_maps}' target='_blank'>Abrir no Google Maps</a>",
                        max_width=300
                    ),
                    icon=folium.Icon(icon="cloud", color="green")
                ).add_to(mapa)
            except Exception as e:
                print(f"Erro ao adicionar marcador: {e}")
                
        folium.Marker(
            [-22.832113794507347, -43.346853465267536],
            popup="Araçatuba Material",
            icon=folium.Icon(icon="home", color="orange")
        ).add_to(mapa)

        st_folium(mapa, width=1800, height=800)
        
        
    def main(self, lista):
        
        self.lista_abas = lista
        lista_navegação = []
        
        for pagina_selecionada in self.lista_abas:
            if pagina_selecionada == "Cadastro de Empresa":
                lista_navegação.append(st.Page(self.cadastro_empresa, title= "🏢 Cadastro de Empresa"))
            elif pagina_selecionada == "Cadastrar novo pedido":
                lista_navegação.append(st.Page(self.cadastro_novo_pedido, title="📝 Cadastrar novo pedido"))
            elif pagina_selecionada == "Consulta de Pedidos":
                lista_navegação.append(st.Page(self.Consulta_pedidos, title="📑 Consulta de Pedidos"))
            elif pagina_selecionada == "Controle de Coleta":
                lista_navegação.append(st.Page(self.Controle_coleta, title="📦 Controle de Coleta"))
            elif pagina_selecionada == "Separar pedido":
                lista_navegação.append(st.Page(self.Separar_pedido, title="📋 Separar pedido"))
            elif pagina_selecionada == "Estoque":
                lista_navegação.append(st.Page(self.Estoque, title="📊 Estoque"))
            elif pagina_selecionada == "Dashboard":
                lista_navegação.append(st.Page(self.Dashboard, title="📈 Dashboard"))
            elif pagina_selecionada == "Rotas":
                lista_navegação.append(st.Page(self.Rotas, title="🚚 Rotas"))
            elif pagina_selecionada == "Rotas clientes":
                lista_navegação.append(st.Page(self.Rotas2, title="🚚 Rotas clientes"))
        
        pg = st.navigation({"Aracatuba parafusos":lista_navegação}, position="sidebar")
        pg.run()



    @st.cache_data
    def Buscar_dados(_self):
        df = pd.read_excel("Mapa novas lojas.xlsx")
        return df

    @st.experimental_dialog("Aviso de pedido")
    def Aviso_pedido(self, aviso):
        st.write(f"{aviso}")
        if st.button("Confirmar"):
            st.rerun()
            
    

    @st.cache_data
    def buscar_empresa(_self, empresa):
        print(f"Função chamada {empresa}")
        return AWS().buscar_cadastro_empresa(empresa)
    
    @st.cache_data
    def buscar_pedidos_nao_pagos(_self):
        print("Pedidos não pagos")
        return [item['S'] for item in AWS().buscar_pedido_nao_pago()]

    @st.cache_data
    def buscar_cadastro_empresas(_self):
        Cadastro_empresas = AWS().Buscar_todos_cadastro_clientes()
        return Cadastro_empresas
    
    @st.cache_data
    def tabela_estoque_formatada(_self, Estoque):
        estoque_atual = AWS().buscar_Estoque(Estoque)
        df = pd.DataFrame(estoque_atual)
        df = df.reset_index()
        df2 = pd.DataFrame()
        df2["Cartela"] = df["index"]           
        df2["Quantidade"] = df["Quantidade"]
        df2['Número'] = df2["Cartela"].str.extract('(\d+)').astype(int)
        df2 = df2.sort_values(by='Número').drop(columns='Número')
        #print("Print_tabela chamada")
        return df2
    
    @st.cache_data
    def tabela_pedido_formatada(_self, Estoque):
        estoque_atual = AWS().buscar_Estoque(Estoque)
        df = pd.DataFrame(estoque_atual)
        df = df.reset_index()
        df2 = pd.DataFrame()
        df2["Cartela"] = df["index"]           
        df2["Quantidade"] = df["Quantidade"]
        df2['Número'] = df2["Cartela"].str.extract('(\d+)').astype(int)
        df2 = df2.sort_values(by='Número').drop(columns='Número')
        #print("Print_tabela chamada")
        return df2
    
    @st.cache_data
    def buscar_clientes(_self):
        return AWS().buscar_clientes()
    
    @st.cache_data
    def buscar_id_pedidos_lojas(_self, empresa):
        return [pedido["S"] for pedido in AWS().buscar_id_pedidos_lojas(empresa)["L"]]